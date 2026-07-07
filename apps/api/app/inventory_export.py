from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from .inventory_helpers import (
    asset_inventory_status,
    catalog_ref,
    list_inventory_assets_filtered,
)
from .models import Asset, AssetMovement, User
from .time_utils import SAO_PAULO, ensure_utc, utc_now

INVENTORY_STATUS_LABELS = {
    "stock": "Estoque",
    "allocated": "Alocado",
    "maintenance": "Em manutenção",
    "retired": "Baixado",
}

LEGACY_TYPE_LABELS = {
    "computer": "Computador",
    "notebook": "Notebook",
    "monitor": "Monitor",
    "printer": "Impressora",
    "network": "Rede",
}

MOVEMENT_ACTION_LABELS = {
    "created": "Cadastro",
    "updated": "Atualização",
    "allocated": "Envio/Alocação",
    "responsible_changed": "Troca de responsável",
    "returned_to_stock": "Devolução ao estoque",
    "maintenance": "Manutenção",
}

EMPTY_VALUE = "Não informado"

EXPORT_HEADERS = (
    "Fornecedor",
    "Tipo",
    "Fabricante",
    "Modelo",
    "Número de série",
    "Setor",
    "Responsável",
    "Situação",
    "Data de recebimento",
    "Data de entrega",
    "Última movimentação",
    "Observações",
)


def export_filename(reference: datetime | None = None) -> str:
    current = (reference or utc_now()).astimezone(SAO_PAULO)
    return f"inventario_adcetei_{current:%Y-%m-%d}.xlsx"


def _catalog_name(ref) -> str:
    payload = catalog_ref(ref)
    return payload["name"] if payload else ""


def _text(value: str | None, *, empty: str = EMPTY_VALUE) -> str:
    cleaned = (value or "").strip()
    return cleaned or empty


def _type_label(asset: Asset) -> str:
    catalog = _catalog_name(asset.equipment_type)
    if catalog:
        return catalog
    legacy = (asset.asset_type or "").strip()
    return LEGACY_TYPE_LABELS.get(legacy, legacy) or EMPTY_VALUE


def _format_date(value: datetime | None) -> str:
    normalized = ensure_utc(value)
    if normalized is None:
        return ""
    return normalized.astimezone(SAO_PAULO).strftime("%d/%m/%Y")


def _format_datetime(value: datetime | None) -> str:
    normalized = ensure_utc(value)
    if normalized is None:
        return ""
    return normalized.astimezone(SAO_PAULO).strftime("%d/%m/%Y %H:%M")


def latest_movements(db: Session, asset_ids: list[int]) -> dict[int, AssetMovement]:
    if not asset_ids:
        return {}
    # ponytail: movement_date + id (not max(id)) — backdated entries must win over newer IDs
    row_number = (
        func.row_number()
        .over(
            partition_by=AssetMovement.asset_id,
            order_by=(AssetMovement.movement_date.desc(), AssetMovement.id.desc()),
        )
        .label("rn")
    )
    ranked = (
        select(AssetMovement.id.label("movement_id"), row_number)
        .where(AssetMovement.asset_id.in_(asset_ids))
        .subquery()
    )
    movement_ids = [
        movement_id
        for movement_id in db.scalars(select(ranked.c.movement_id).where(ranked.c.rn == 1))
        if movement_id
    ]
    if not movement_ids:
        return {}
    movements = db.scalars(select(AssetMovement).where(AssetMovement.id.in_(movement_ids)))
    return {movement.asset_id: movement for movement in movements}


def _movement_summary(movement: AssetMovement | None) -> str:
    if movement is None:
        return EMPTY_VALUE
    when = _format_datetime(movement.movement_date or movement.created_at)
    action = MOVEMENT_ACTION_LABELS.get(movement.action, movement.action)
    return f"{when} · {action}" if when else action


def asset_export_row(asset: Asset, last_movement: AssetMovement | None) -> tuple[Any, ...]:
    status = asset_inventory_status(asset)
    return (
        _text(_catalog_name(asset.supplier)),
        _type_label(asset),
        _text(_catalog_name(asset.manufacturer_ref) or asset.manufacturer),
        _text(_catalog_name(asset.equipment_model) or asset.model),
        _text(asset.serial_number, empty="Não informado"),
        _text(_catalog_name(asset.sector) or asset.location),
        _text(asset.assigned_user.full_name if asset.assigned_user else None),
        INVENTORY_STATUS_LABELS.get(status, status),
        _format_date(asset.received_at) or EMPTY_VALUE,
        _format_date(asset.delivered_at) or EMPTY_VALUE,
        _movement_summary(last_movement),
        _text(asset.notes),
    )


def build_inventory_export_workbook(
    assets: list[Asset],
    *,
    last_movements: dict[int, AssetMovement],
    exported_by: User,
    exported_at: datetime | None = None,
) -> bytes:
    exported_at = ensure_utc(exported_at) or utc_now()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Inventário"

    header_fill = PatternFill("solid", fgColor="D9E8F5")
    header_font = Font(bold=True, color="1A2332")
    meta_font = Font(bold=True, color="164F84")

    sheet["A1"] = "Portal Interno ADCETEI"
    sheet["A1"].font = meta_font
    sheet["A2"] = "Exportação de Inventário"
    sheet["A3"] = (
        f"Exportado em {_format_datetime(exported_at)} por {exported_by.full_name} "
        f"({exported_by.email})"
    )
    sheet["A4"] = f"Registros exportados: {len(assets)}"

    header_row = 6
    for column, title in enumerate(EXPORT_HEADERS, start=1):
        cell = sheet.cell(row=header_row, column=column, value=title)
        cell.font = header_font
        cell.fill = header_fill

    data_start = header_row + 1
    for offset, asset in enumerate(assets):
        row_index = data_start + offset
        for column, value in enumerate(asset_export_row(asset, last_movements.get(asset.id)), start=1):
            sheet.cell(row=row_index, column=column, value=value)

    last_row = max(data_start, data_start + len(assets) - 1)
    last_column = get_column_letter(len(EXPORT_HEADERS))
    sheet.auto_filter.ref = f"A{header_row}:{last_column}{last_row}"
    sheet.freeze_panes = f"A{data_start}"

    column_widths = (22, 18, 18, 24, 22, 22, 28, 16, 20, 18, 28, 36)
    for index, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_inventory_assets(
    db: Session,
    *,
    actor: User,
    status_filter: str | None,
    equipment_type_id: int | None,
    sector_id: int | None,
    search: str | None,
) -> tuple[bytes, str, int]:
    assets = list_inventory_assets_filtered(
        db,
        status_filter=status_filter,
        equipment_type_id=equipment_type_id,
        sector_id=sector_id,
        search=search,
    )
    last_movements = latest_movements(db, [asset.id for asset in assets])
    content = build_inventory_export_workbook(
        assets,
        last_movements=last_movements,
        exported_by=actor,
    )
    return content, export_filename(), len(assets)
